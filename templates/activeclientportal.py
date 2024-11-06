import os
import dash
from dash import dcc, html, Input, Output, State, exceptions
import dash_bootstrap_components as dbc
import pandas as pd
import plotly.express as px
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

# Initialize the Dash app with Bootstrap components for styling
app = dash.Dash(__name__, external_stylesheets=[dbc.themes.BOOTSTRAP])

# Define the valid clients dictionary
valid_clients = {
    "EI": "EI2024!",
    "AL": "A&L2024!",
    "DLI": "DLI2024!",
}

# Set up a dictionary to track login attempts for each session
app.clients_data = {"attempts": 0, "authenticated": False}

# Layout of the Dash app
app.layout = dbc.Container([
    dbc.Row([
        dbc.Col([
            # Sidebar layout
            html.H2("Client Authentication", className="text-center"),
            dcc.Input(id="client-id", type="text", placeholder="Client ID", className="mb-2"),
            dcc.Input(id="client-password", type="password", placeholder="Client Password", className="mb-2"),
            html.Button("Submit", id="submit-button", n_clicks=0, className="btn btn-primary"),
            html.Div(id="login-status", className="mt-3"),
        ], width=3, className="sidebar"),
        
        # Main content area
        dbc.Col([
            html.H2("Financial Forecast Model", className="text-center"),
            html.Div(id="tab-container"),
            dcc.Tabs(id="tabs", value=None),
            html.Div(id="tab-content", className="mt-3"),
        ], width=9)
    ])
], fluid=True)

# Helper function to authenticate user
def authenticate(client_id, client_password):
    if client_id not in valid_clients:
        return "Client ID not found"
    elif valid_clients[client_id] != client_password:
        return "Incorrect password"
    else:
        return "Authenticated"

# Callback to handle the login attempts and authentication
@app.callback(
    Output("login-status", "children"),
    Output("tab-container", "children"),
    Input("submit-button", "n_clicks"),
    State("client-id", "value"),
    State("client-password", "value"),
    prevent_initial_call=True
)
def login(n_clicks, client_id, client_password):
    if n_clicks > 0:
        # Check if already authenticated
        if app.clients_data["authenticated"]:
            raise exceptions.PreventUpdate

        # Track attempts
        app.clients_data["attempts"] += 1
        
        # Check if attempts exceeded
        if app.clients_data["attempts"] > 5:
            return "Maximum login attempts exceeded.", None

        # Attempt to authenticate
        auth_status = authenticate(client_id, client_password)
        
        if auth_status == "Authenticated":
            app.clients_data["authenticated"] = True
            folder_path = r"C:\Users\jorda\OneDrive\Documents\GitHub\streamlit\Website"  # Replace with actual folder path
            file_name = f"{client_id}_FFM.xlsx"
            file_path = os.path.join(folder_path, file_name)

            # Try to access the file with the password
            if os.path.exists(file_path):
                try:
                    # Open workbook and retrieve sheet names
                    workbook = load_workbook(filename=file_path, read_only=True, keep_vba=True)
                    sheet_names = workbook.sheetnames

                    # Create tabs for each sheet
                    tabs = dcc.Tabs(
                        id="tabs",
                        value=sheet_names[0],
                        children=[dcc.Tab(label=sheet, value=sheet) for sheet in sheet_names]
                    )

                    return "Login successful.", tabs

                except InvalidFileException:
                    return "Unable to open the file. The file may be corrupt or inaccessible.", None
                except KeyError:
                    return "Incorrect password for this file.", None
            else:
                return f"No financial forecast model found for Client ID '{client_id}'.", None
        else:
            return auth_status, None

    return "", None

# Callback to display selected sheet content
@app.callback(
    Output("tab-content", "children"),
    Input("tabs", "value"),
    State("client-id", "value"),
    State("client-password", "value"),
    prevent_initial_call=True
)
def display_sheet_content(sheet_name, client_id, client_password):
    if sheet_name:
        folder_path = r"C:\Users\jorda\OneDrive\Documents\GitHub\streamlit\Website"  # Replace with actual folder path
        file_name = f"{client_id}_FFM.xlsx"
        file_path = os.path.join(folder_path, file_name)
        
        # Load the sheet as a DataFrame
        sheet_data = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')

        # Drop any rows or columns that are completely empty
        sheet_data.dropna(how="all", inplace=True)

        # Format table for nicer display
        table = dbc.Table.from_dataframe(sheet_data, striped=True, bordered=True, hover=True)

        # Create a sample graph for demonstration (based on the first two numeric columns if available)
        numeric_columns = sheet_data.select_dtypes(include="number").columns
        if len(numeric_columns) >= 2:
            fig = px.line(sheet_data, x=numeric_columns[0], y=numeric_columns[1],
                          title=f"{sheet_name} Data Overview")
            graph = dcc.Graph(figure=fig)
        else:
            graph = html.P("Not enough data for visualization")

        return html.Div([table, html.Br(), graph])

# Run the Dash app
if __name__ == "__main__":
    app.run_server(debug=True)
